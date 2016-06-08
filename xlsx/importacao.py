from decimal import Decimal

from openpyxl import load_workbook

from xlsx.models import Venda


class ImportacaoXlsx():
    @staticmethod
    def importar_vendas(obj):
        """
        Recebe o path do arquivo enviado

        :param arquivo: path do arquivo
        """

        pasta = load_workbook(filename=obj.arquivo.path, use_iterators=True)
        name_plan = pasta.get_sheet_names()[0]  # Selecionado a primeira planilha da pasta de trabalho
        planilha = pasta.get_sheet_by_name(name_plan)

        # Iniciando loop em todos os registros da planilha
        for idx, row in enumerate(planilha.iter_rows()):
            if idx != 0:
                cl1 = row[0]  # nota
                cl2 = row[1]  # valor
                cl3 = row[2]  # vencimento

                data = {
                    'nota': str(cl1.value),
                    'valor': Decimal(cl2.value),
                    'vencimento': str(cl3.value)[:10],
                    'importacao': obj,
                }

                if cl1.value is not None:  # Validando para não incluir registros em branco
                    Venda.objects.import_venda(data)  # Inserindo os dados
